import random
from random import randint

import names
from openpyxl.workbook import Workbook

from random_start_time import *
from random_end_time import *

wb = Workbook()
ws = wb.active
ws.title = "DataSet"

randint(1000, 10000)
city = ["Delhi", "Mumbai", "Pune", "Kolkata", "Chennai", "Bangalore", "Jaipur", "Jaipur"]

product_name = ["Software 1", "Software 2", "Software 3", "Software 4", "Software 5", "Software 6", "Software 7",
                "Software 8", "Software 9", "Software 10"]

feature_name = ["Feature A", "Feature B", "Feature C", "Feature D", "Feature E", "Feature F", "Feature G", "Feature H",
                "Feature I", "Feature J", "Feature K", "Feature L", "Feature M"]


ws['A1'] = "Customer_Name"
ws['B1'] = "Product_ID"
ws['C1'] = "Product_Name"
ws['D1'] = "Feature_ID"
ws['E1'] = "Feature_Name"
ws['F1'] = "Invoice_Date"
ws['G1'] = "Cost"
ws['H1'] = "City"
ws['I1'] = "Contact_Number"
ws['J1'] = "Partner_Name"

starting_row = 2
for i in range(1000):
    name = names.get_first_name()
    name_2 = names.get_first_name()

    ws.cell(row=starting_row, column=1).value = name
    ws.cell(row=starting_row, column=2).value = x = randrange(0, 5)
    ws.cell(row=starting_row, column=3).value = product_name[x]
    ws.cell(row=starting_row, column=4).value = y = randrange(0, 9)
    ws.cell(row=starting_row, column=5).value = feature_name[y]
    ws.cell(row=starting_row, column=6).value = random_date(e1, e2)
    ws.cell(row=starting_row, column=7).value = randint(800, 2000)
    ws.cell(row=starting_row, column=8).value = city[randrange(0, 8)]
    ws.cell(row=starting_row, column=9).value = randint(9100000000, 9999999999)
    ws.cell(row=starting_row, column=10).value = name_2

    starting_row = starting_row + 1

wb.save('Data.xlsx')
